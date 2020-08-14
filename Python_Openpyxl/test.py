import openpyxl
from openpyxl import Workbook

# 实例化
# wb = Workbook()
# 激活 worksheet
# ws = wb.active

# name = input('请输入这个Excel表格的名字:')
# wb.save('{}.xlsx'.format(name))

from openpyxl import load_workbook

# wb2 = load_workbook('你没有取名字哟.xlsx')


# sheet = wb2['Sheet1']
# for row in sheet.rows:
#
#     # 循环遍历每一个单元格
#     for cell in row:
#         # 获取单元格的内容
#         print(cell.value, end=',')
#     print()
# data = [[' BOOK', 50, 3], ['APPLE', 100, 1], ['BANANA', 200, 0.5]]
#
#
# def write_to_excel_list(data):
#     wb = load_workbook('你没有取名字哟.xlsx')
#     # ws = wb.active
#     sheet = wb['Sheet1']
#     for row, item in enumerate(data):  # 0 [' BOOK', 50, 3]
#         for column, cellValue in enumerate(item):  # 0 ' BOOK'
#             sheet.cell(row=row + 1, column=column + 1, value=cellValue)
#     wb.save(filename='你没有取名字哟.xlsx')
#
#
# write_to_excel_list(data=data)

# from Python_Openpyxl.python_for_excel import TestExcel
# path = 'C:/Users/YAE/Desktop/python_all/Python_Openpyxl/temp2'
# a = TestExcel()
# a.write_to_excel_json(path)

# import json
# import pandas as pd
#
# data = []  # 用于存储每一行的Json数据
# with open('C:/Users/YAE/Desktop/python_all/Python_Openpyxl/temp2', 'r', encoding='UTF-8') as fr:
#     for line in fr:
#         j = json.loads(line)
#         data.append(j)
#
# df = pd.DataFrame()  # 最后转换得到的结果
# for line in data:
#     for i in line:
#         df1 = pd.DataFrame([i])
#         df = df.append(df1)
#
# # 在excel表格的第1列写入, 不写入index
# df.to_excel('data2.xlsx', sheet_name='Data', startcol=0, index=False)
from Python_Openpyxl.python_for_excel import TestExcel

a = TestExcel()
a.show_all()