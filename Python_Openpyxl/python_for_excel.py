import json
import pandas as pd

import openpyxl
from openpyxl import load_workbook


class TestExcel:

    def __init__(self):
        # name = input('输入你要打开文件的名称:')
        wb = load_workbook('{}.xlsx'.format('data2'))
        sheet = wb['Data']
        self.wb = wb
        self.sheet = sheet

    # 显示表格的所有内容
    def show_all(self):
        for row in self.sheet.rows:
            # 循环遍历每一个单元格
            for cell in row:
                # 获取单元格的内容
                print('%-24s' % cell.value, end='')
            print()

    # 创建一个新的Excel表格
    @staticmethod
    def create_new_excel(excelname='你没有取名字哟', title='Sheet1'):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
        wb.save("{}.xlsx".format(excelname))

    # 将列表内容写入Excel中
    @staticmethod
    def write_to_excel_list(data):
        wb = load_workbook('你没有取名字哟.xlsx')
        sheet = wb['Sheet1']
        for row, item in enumerate(data):
            for column, cellValue in enumerate(item):
                sheet.cell(row=row + 1, column=column + 1, value=cellValue)
        wb.save(filename='你没有取名字哟.xlsx')

    @staticmethod
    def write_to_excel_json(path):
        data = []  # 用于存储每一行的Json数据
        with open(path, 'r', encoding='utf-8') as fr:
            for line in fr:
                j = json.loads(line)
                data.append(j)

        df = pd.DataFrame()  # 最后转换得到的结果
        for line in data:
            for i in line:
                df1 = pd.DataFrame([i])
                df = df.append(df1)

        df.to_excel('data.xlsx', sheet_name='Data', startcol=0, index=False)
