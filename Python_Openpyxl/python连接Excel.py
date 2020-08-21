import openpyxl

wb = openpyxl.load_workbook('test.xlsx')  # 打开Book1文档，实力化出一个wb对象

print(wb.sheetnames)  # 获取当前工作簿里的所有表
print(wb.active)  # 正在使用的表

sheet = wb['Sheet1']
print(sheet.title)  # 获取工作表的名称

print(sheet.cell(row=2, column=3).value)  # 获取指定位置的信息

cell = sheet['B1']  # 指定单元格，对单元格的相关进行操作
print(cell.row, cell.column, cell.value)

print(sheet.max_row)
print(sheet.max_column)  # 获取最大行数和最大列数


##访问所有单元格的信息
print(type(sheet.rows))  # sheet的所有行和所有列都是一个生成器
for row in sheet.rows:
    for info in row:
        print(info.value)

# wb.save(filename='test.xlsx')  # 保存所有修改信息
